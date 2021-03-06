from openpyxl import Workbook
from openpyxl.styles import Side, Border
from openpyxl.utils.cell import coordinate_from_string, get_column_letter


def make_border(start, end):
    bd_thin = Side(style='thin')
    bd_thick = Side(style='medium')
    
    # 左上为起、右下为止的坐标
    start_col = coordinate_from_string(start)[0]
    start_row = coordinate_from_string(start)[1]
    end_col = coordinate_from_string(end)[0]
    end_row = coordinate_from_string(end)[1]
    
    start_pos = "".join(map(str, tuple(start)))
    end_pos = "".join(map(str, tuple(end)))

    for row in ws[start_pos:end_pos]:
        for cell in row:
            # 当前单元格的坐标
            cell.border = Border(left=bd_thin, top=bd_thin, right=bd_thin, bottom=bd_thin)
            i_col = get_column_letter(cell.column)
            i_row = cell.row
            print(i_col, i_row)
            
            if i_col == start_col and i_col == end_col and i_row == start_row and i_row == end_row:
                cell.border = Border(left=bd_thick, right=bd_thick, top=bd_thick, bottom=bd_thick)
            elif i_col == start_col and i_col == end_col and i_row == start_row:  # 竖列的上
                cell.border = Border(left=bd_thick, right=bd_thick, top=bd_thick, bottom=bd_thin)
            elif i_col == start_col and i_col == end_col and i_row == end_row:  # 竖列的下
                cell.border = Border(left=bd_thick, right=bd_thick, top=bd_thin, bottom=bd_thick)
            elif i_col == start_col and i_row == start_row and i_row == end_row:  # 横列的左
                cell.border = Border(left=bd_thick, right=bd_thin, top=bd_thick, bottom=bd_thick)
            elif i_col == end_col and i_row == start_row and i_row == end_row:  # 横列的右
                cell.border = Border(left=bd_thin, right=bd_thick, top=bd_thick, bottom=bd_thick)
            elif i_col == start_col and i_col == end_col:  # 竖列的两边
                cell.border = Border(left=bd_thick, right=bd_thick, top=bd_thin, bottom=bd_thin)
            elif i_row == start_row and i_row == end_row:  # 横排的上下
                cell.border = Border(left=bd_thin, right=bd_thin, top=bd_thick, bottom=bd_thick)
            elif i_col == start_col and i_row == start_row:  # 左上角
                cell.border = Border(left=bd_thick, right=bd_thin, top=bd_thick, bottom=bd_thin)
            elif i_col == end_col and i_row == end_row:  # 右下角
                cell.border = Border(left=bd_thin, right=bd_thick, top=bd_thin, bottom=bd_thick)
            elif i_col == start_col and i_row == end_row:  # 左下角
                cell.border = Border(left=bd_thick, right=bd_thin, top=bd_thin, bottom=bd_thick)
            elif i_col == end_col and i_row == start_row:  # 右上角
                cell.border = Border(left=bd_thin, right=bd_thick, top=bd_thick, bottom=bd_thin)
            elif i_col == start_col:  # 左边
                cell.border = Border(left=bd_thick, right=bd_thin, top=bd_thin, bottom=bd_thin)
            elif i_col == end_col:  # 右边
                cell.border = Border(left=bd_thin, right=bd_thick, top=bd_thin, bottom=bd_thin)
            elif i_row == start_row:  # 上边
                cell.border = Border(left=bd_thin, right=bd_thin, top=bd_thick, bottom=bd_thin)
            elif i_row == end_row:  # 下边
                cell.border = Border(left=bd_thin, right=bd_thin, top=bd_thin, bottom=bd_thick)
            else:
                cell.border = Border(left=bd_thin, right=bd_thin, top=bd_thin, bottom=bd_thin)
    print('-'*20)


if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active

    # --------示例--------
    # 多维
    make_border('B3', 'E5')
    # 单个字符
    make_border('B7', 'B7')
    # 一竖列
    make_border('B9', 'B12')
    # 一横排
    make_border('B14', 'E14')

    wb.save("sample.xlsx")

